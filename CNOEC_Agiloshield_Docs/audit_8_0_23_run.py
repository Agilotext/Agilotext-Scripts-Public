#!/usr/bin/env python3
"""
Audit ultra 8.0.23 — lot 260305_1209_8.0.23
Extrait texte des PDF/Excel anonymisés, cherche fuites.
"""

import re
import subprocess
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

OUT_DIR = Path("/Users/florianbauer/Downloads/260305_1209_8.0.23/OUT")
IN_DIR = Path("/Users/florianbauer/Downloads/260305_1209_8.0.23/IN")
REPORT_PATH = Path("/Users/florianbauer/Documents/AGILOTEXT/Agilotext-Scripts-Public/CNOEC_Agiloshield_Docs/AUDIT_ULTRA_8.0.23_260305_1209.md")

# P0 = critique, P1 = important, P2 = modéré
LEAK_PATTERNS = {
    "P0_IBAN": re.compile(
        r"\bFR\d{2}[\s.]?\d{4}[\s.]?\d{4}[\s.]?\d{4}[\s.]?\d{4}[\s.]?\d{4}[\s.]?\d{3}\b",
        re.IGNORECASE,
    ),
    "P0_RIB": re.compile(
        r"\b\d{5}[\s.\-]*\d{5}[\s.\-]*[A-Z0-9]{11}[\s.\-]*\d{2}\b",
        re.IGNORECASE,
    ),
    "P0_NIR": re.compile(
        r"\b[12][\s.\-]*\d{2}[\s.\-]*\d{2}[\s.\-]*\d{2}[\s.\-]*\d{3}[\s.\-]*\d{3}[\s.\-]*\d{2}\b",
        re.IGNORECASE,
    ),
    "P1_SIRET": re.compile(
        r"\d{3}[\s.]?\d{3}[\s.]?\d{3}[\s.]?\d{5}",
        re.IGNORECASE,
    ),
    "P1_SIREN": re.compile(
        r"\b\d{3}[\s.]?\d{3}[\s.]?\d{3}\b",
        re.IGNORECASE,
    ),
    "P1_TVA": re.compile(
        r"\bFR\s*[0-9A-Z]{2}[\s.]?\d{9}\b",
        re.IGNORECASE,
    ),
    # TVA UE (8.0.26) — aligné anon_replacer.py
    "P1_TVA_AT": re.compile(r"(?<!\w)AT\s*U?\s*\d{8}(?!\w)", re.IGNORECASE),
    "P1_TVA_DE": re.compile(r"(?<!\w)DE\s*\d{9}(?!\w)", re.IGNORECASE),
    "P1_TVA_IT": re.compile(r"(?<!\w)IT\s*\d{11}(?!\w)", re.IGNORECASE),
    "P1_TVA_ES": re.compile(r"(?<!\w)(?:ES|E)\s*[A-Z0-9]\d{7}[A-Z0-9](?!\w)", re.IGNORECASE),
    "P1_TVA_NL": re.compile(r"(?<!\w)NL\s*\d{9}B\d{2}(?!\w)", re.IGNORECASE),
    "P1_TVA_BE": re.compile(r"(?<!\w)BE\s*0?\d{9}(?!\w)", re.IGNORECASE),
    "P1_TVA_LU": re.compile(r"(?<!\w)LU\s*\d{8}(?!\w)", re.IGNORECASE),
    # PDL EDF (8.0.26)
    "P1_PDL": re.compile(
        r"(?i)(?:PDL|point\s+de\s+livraison)\s*[:\s]*\d{14}(?!\d)",
    ),
    # CompAuxNum FEC (8.0.26)
    "P1_CompAuxNum": re.compile(r"\b(CC|CE|CF)[A-Z][A-Z0-9]{7,}\b"),
    "P2_NOMS": re.compile(
        r"\b(DURAND|FONTAINE|BERNARD|PETIT|RODRIGUEZ|MOREAU|DUPONT|LAMBERT|MARTIN)\b",
        re.IGNORECASE,
    ),
    "P2_LIEUX": re.compile(
        r"\b(MARSEILLE|LYON|PARIS|MONTPELLIER|TOULOUSE|NICE|NANTES|STRASBOURG|BORDEAUX|LILLE|RENNES)\b",
        re.IGNORECASE,
    ),
    "EMAIL": re.compile(
        r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",
        re.IGNORECASE,
    ),
    "PHONE": re.compile(
        r"\b0[1-9][\s.]?\d{2}[\s.]?\d{2}[\s.]?\d{2}[\s.]?\d{2}\b|\+\s*33[\s.]?\d[\s.]?\d{2}[\s.]?\d{2}[\s.]?\d{2}[\s.]?\d{2}\b",
        re.IGNORECASE,
    ),
}

# Placeholders attendus = données bien masquées
PLACEHOLDER_OK = re.compile(r"\[(?:IBAN|RIB|NIR|SIRET|SIREN|VAT|PDL|FEC_AUX|EMAIL|PHONE|PERSON|LOCATION|ORGANIZATION)\]")


def extract_text_pdf(path: Path) -> str:
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode == 0:
            return result.stdout or ""
    except Exception as e:
        return f"[ERREUR: {e}]"
    return ""


def extract_text_excel(path: Path) -> str:
    if load_workbook is None:
        return "[ERREUR: openpyxl non installé]"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        chunks = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        chunks.append(str(cell.value))
        wb.close()
        return "\n".join(chunks)
    except Exception as e:
        return f"[ERREUR: {e}]"


def search_leaks(text: str) -> dict:
    findings = defaultdict(list)
    for leak_type, pattern in LEAK_PATTERNS.items():
        for m in pattern.finditer(text):
            # Exclure si c'est dans un placeholder type [SIRET] ou contexte masqué
            raw = m.group(0)
            start = max(0, m.start() - 50)
            end = min(len(text), m.end() + 50)
            ctx = text[start:end].replace("\n", " ")
            if "[SIRET]" in ctx or "[SIREN]" in ctx or "[IBAN]" in ctx or "[NIR]" in ctx or "[VAT]" in ctx or "[PDL]" in ctx or "[FEC_AUX]" in ctx:
                continue  # faux positif: le placeholder est à côté
            findings[leak_type].append((raw, ctx))
    return dict(findings)


def main():
    # Documents à auditer
    pdf_anon = sorted(OUT_DIR.glob("*.ANON.pdf"))
    excel_anon = [p for p in OUT_DIR.iterdir() if p.is_file() and "ANON" in p.name and (".xlsx" in p.name or ".xls" in p.name)]

    all_docs = []
    doc_findings = {}
    type_totals = defaultdict(int)
    examples = {}

    for p in pdf_anon:
        text = extract_text_pdf(p)
        findings = search_leaks(text)
        all_docs.append((p.name, "PDF", findings))
        doc_findings[p.name] = findings
        for lt, matches in findings.items():
            type_totals[lt] += len(matches)
        ex = {lt: [(m, c[:100]) for m, c in matches[:2]] for lt, matches in findings.items()}
        examples[p.name] = ex

    for p in excel_anon:
        text = extract_text_excel(p)
        findings = search_leaks(text)
        all_docs.append((p.name, "Excel", findings))
        doc_findings[p.name] = findings
        for lt, matches in findings.items():
            type_totals[lt] += len(matches)
        ex = {lt: [(m, c[:100]) for m, c in matches[:2]] for lt, matches in findings.items()}
        examples[p.name] = ex

    total = len(all_docs)
    sans_faille = [d[0] for d in all_docs if not doc_findings[d[0]]]
    avec_faille = [d[0] for d in all_docs if doc_findings[d[0]]]

    # Score /20 : déduction par type de faille
    p0_count = type_totals.get("P0_IBAN", 0) + type_totals.get("P0_RIB", 0) + type_totals.get("P0_NIR", 0)
    p1_count = (
        type_totals.get("P1_SIRET", 0) + type_totals.get("P1_SIREN", 0) + type_totals.get("P1_TVA", 0)
        + type_totals.get("P1_TVA_AT", 0) + type_totals.get("P1_TVA_DE", 0) + type_totals.get("P1_TVA_IT", 0)
        + type_totals.get("P1_TVA_ES", 0) + type_totals.get("P1_TVA_NL", 0) + type_totals.get("P1_TVA_BE", 0)
        + type_totals.get("P1_TVA_LU", 0) + type_totals.get("P1_PDL", 0) + type_totals.get("P1_CompAuxNum", 0)
    )
    p2_count = type_totals.get("P2_NOMS", 0) + type_totals.get("P2_LIEUX", 0)

    # Grille : 20 - (P0*2) - (P1*0.5) - (P2*0.1), plancher 0
    note = 20.0
    note -= p0_count * 2.0
    note -= p1_count * 0.5
    note -= p2_count * 0.1
    note -= (type_totals.get("EMAIL", 0) + type_totals.get("PHONE", 0)) * 0.3
    note = max(0, min(20, note))

    # Complétude IN vs OUT
    in_pdfs = list(IN_DIR.rglob("*.pdf"))
    in_xlsx = list(IN_DIR.rglob("*.xlsx"))
    in_count = len([p for p in in_pdfs if p.suffix.lower() == ".pdf"]) + len(in_xlsx)
    out_count = len(pdf_anon) + len(excel_anon)

    lines = [
        "# Audit ultra 8.0.23 — lot 260305_1209",
        "",
        f"**Date** : {datetime.now().strftime('%d %B %Y %H:%M')}",
        f"**Répertoire** : `{OUT_DIR}`",
        "",
        "---",
        "",
        "## 1. Complétude documentsIn / documentsOut",
        "",
        f"- **IN** : {in_count} documents (PDF + Excel)",
        f"- **OUT** : {out_count} documents anonymisés",
        f"- **Écart** : {in_count - out_count}",
        "",
        "---",
        "",
        "## 2. Verdict global",
        "",
        f"- **Documents sans fuite** : {len(sans_faille)} / {total}",
        f"- **Documents avec fuite** : {len(avec_faille)} / {total}",
        f"- **P0 (IBAN/RIB/NIR)** : {p0_count}",
        f"- **P1 (SIRET/SIREN/TVA)** : {p1_count}",
        f"- **P2 (Noms/Lieux)** : {p2_count}",
        "",
        f"### Note : **{note:.1f} / 20**",
        "",
        "---",
        "",
        "## 3. Fuites par type",
        "",
        "| Type | Occurrences |",
        "|------|-------------|",
    ]
    for lt in sorted(LEAK_PATTERNS.keys()):
        lines.append(f"| {lt} | {type_totals[lt]} |")
    lines.append("")
    lines.extend(["---", "", "## 4. Documents sans faille", ""])
    for n in sans_faille[:30]:
        lines.append(f"- `{n}`")
    if len(sans_faille) > 30:
        lines.append(f"- ... et {len(sans_faille) - 30} autres")
    lines.append("")
    lines.extend(["---", "", "## 5. Documents avec fuites (détail)", ""])
    for n in sorted(avec_faille):
        f = doc_findings[n]
        counts = " | ".join(f"{k}: {len(v)}" for k, v in f.items() if v)
        lines.append(f"### {n}")
        lines.append(f"**{counts}**")
        lines.append("")
        for lt, matches in sorted(f.items()):
            if not matches:
                continue
            for m, ctx in matches[:2]:
                ctx_trim = (ctx.strip()[:90] + "...") if len(ctx.strip()) > 90 else ctx.strip()
                lines.append(f"- **{lt}** : `{m}` → *{ctx_trim}*")
        lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Rapport : {REPORT_PATH}")
    print(f"Note : {note:.1f}/20 | P0={p0_count} | P1={p1_count} | P2={p2_count}")
    print(f"Sans faille : {len(sans_faille)} | Avec faille : {len(avec_faille)}")


if __name__ == "__main__":
    main()
